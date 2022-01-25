import openpyxl as xl

wb = xl.load_workbook('business_report.xlsx')
sheet = wb['Sheet1']

head = sheet.cell(1, 5)
head.value = 'Profit Percent'
for row in range(2, 5):
    profit_cell = sheet.cell(row, 3)
    cost_cell = sheet.cell(row, 2)
    profit_percent = (profit_cell.value/cost_cell.value) * 100
    cell_for_profit_percent = sheet.cell(row, 5)
    cell_for_profit_percent.value = profit_percent

wb.save('new_business_report.xlsx')


    
    
